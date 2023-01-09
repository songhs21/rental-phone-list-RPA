#_-_encoding=utf8_-_
#__author__="huiseong.song"

from tkinter import *
import tkinter.ttk as ttk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime, timedelta
import os

########################################################################################

#Tkinter 생성
root = Tk() # Tkinter 생성
root.title("IMEI 추가 귀찮은ww") # 타이틀 설정
root.geometry("400x100") # 창 크기 설정
root.resizable(False, False) # 창 크기 변경 가능 여부 설정


# 엑셀 파일 불러오기
today = datetime.today()

# 정합성 대여 단말 리스트 엑셀 불러오기

directory = "D:\Python" # 파일이 존재하는 위치 변수
for filename in os.listdir(directory):
        f = os.path.join(directory, filename) # 파일 리스트 생성
if "정합성 단말 대여 리스트_"+ str(today).replace("-","")[2:8] +".xlsx" in f:
    rental = load_workbook("정합성 단말 대여 리스트_"+ str(today).replace("-","")[2:8] +".xlsx") # today 날짜의 파일이 있으면 파일 불러오기
    print(str(today)[2:10] + " 파일 오픈")
else:
    for n in range(1, 14):
        if "정합성 단말 대여 리스트_"+ str(today-timedelta(days=n)).replace("-","")[2:8] +".xlsx" in f: # today 날짜의 파일이 없는 경우 하루씩 돌아가며 체크
            rental = load_workbook("정합성 단말 대여 리스트_"+ str(today-timedelta(days=n)).replace("-","")[2:8] +".xlsx") # 해당 날짜의 파일 불러오기
            rentalSheet = rental.active # 엑셀 시트 활성화
            print(str(today-timedelta(days=n))[2:10] + " 파일 오픈") # 오픈한 파일의 날짜를 출력.
            break
        else:
            print(str(today-timedelta(days=n))[2:10] + " 파일 없음")

# T4팀 보유 단말 리스트 엑셀 불러오기
db = load_workbook("list.xlsx") 
dbSheet = db.active 

########################################################################################

frame1 = Frame(root) # 프레임 생성
frame1.pack(side="left", fill="both", pady=(18,0), expand=True) # 프레임 표시

# IMEI
Label(frame1, text="IMEI").pack() # "이름" 라벨 생성 후 pack
num = Entry(frame1, width=30) # 텍스트 필드 생성 후 num 에 저장
num.pack() # num pack

# 대여자
lender = ttk.Combobox(frame1, state="readonly", values=["배진우", "오정민", "송희성"], width=8) # readonly = 입력 불가 / disable = 비활성화
lender.current(2) # 초기 표시 데이터 값 지정
lender.pack(pady=(10, 0)) # 

########################################################################################

def Add():

    days = rentalSheet.cell(row=2, column=12).value # K2의 날짜를 체크할 변수
    if days not in str(today): # today의 날짜 값과 days의 날짜 값이 다른지 체크
        rentalSheet.cell(row=2, column=12, value=str(today)[0:11]+" 확인") # 다르면 K2에 추가하는 날의 날짜로 변경
    
    # 단말 대여 처리
    imei = num.get() # 입력 받은 imei를 저장할 변수
    Rimeis = [] # 대여 리스트 엑셀의 imei를 담을 리스트
    DBimeis = [] # 보유 리스트 엑셀의 imei를 담을 리스트

    if len(imei) == 15:

        for dx in dbSheet["E"]:
            DBimeis.append(dx.value) # list 엑셀에서 E열값을 받아와 imei 리스트 생성

        for rx in rentalSheet["F"]:
            Rimeis.append(rx.value) # 대여 리스트 엑셀에서 F열값을 받아와 imei 리스트 생성

        Eloc = (DBimeis.index(int(imei)))+1 # 보유 리스트 엑셀 E열값 중에서 입력받은 IMEI 값의 rows 값 추출

        CPinfo = [] # 입력받은 IMEI의 해당 단말 정보를 저장할 리스트
        for x in range(1, dbSheet.max_column+1):
            CPinfo.append(dbSheet.cell(row=Eloc, column=x).value) # 보유 리스트 엑셀에서 가져온 imei 단말 위치의 행 전체 값을 추출
        
        
        if imei not in Rimeis: # IMEI 값이 IMEIS에 들어있는지 체크

            name = lender.get()
            x= rentalSheet.max_row+1# 대여 리스트에 작성할 행을 저장
            n=0 # 보유 리스트 엑셀에서 추출한 데이터 리스트 순서를 체크할 변수

            for y in range(rentalSheet.max_column): # 
                rentalSheet[x][y].font = Font(name="맑은 고딕", size=10, color="000000")
                rentalSheet[x][y].border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
                rentalSheet[x][y].alignment = Alignment(horizontal="center", vertical="center")


            for y in range(2, 7):
                rentalSheet.cell(row=x, column=y, value=str(CPinfo[n])) # list 엑셀에서 가져온 데이터를 대여 리스트 엑셀 마지막 행에 순차적으로 입력 
                n+=1
            rentalSheet.cell(row=rentalSheet.max_row, column=7, value="정합성/신뢰성")
            rentalSheet.cell(row=rentalSheet.max_row, column=8, value=name)
            rentalSheet.cell(row=rentalSheet.max_row, column=9, value="대여")
            rentalSheet.cell(row=rentalSheet.max_row, column=10, value=str(today)[0:10])
            rentalSheet.cell(row=rentalSheet.max_row, column=11, value="O")
            rentalSheet.cell(row=rentalSheet.max_row, column=12, value="미반납")
            if name == "오정민" or name == "송희성":
                rentalSheet.cell(row=rentalSheet.max_row, column=13, value="MR 검증중")
            elif name == "배진우":
                rentalSheet.cell(row=rentalSheet.max_row, column=13, value="PCT 장비 사용중")

            n=3 # 입력을 시작할 row 값
            for x in range(rentalSheet.max_row+1):
                rentalSheet.cell(row=n, column=1, value=x+1) # 3번째 줄 부터 No값 삽입
                n+=1 # 다음 행으로
                if n > rentalSheet.max_row:
                    break # n 값이 마지막 행을 넘어갈 경우 반복문 종료
            
            rental.save("정합성 단말 대여 리스트_"+ str(today).replace("-","")[2:8] +".xlsx") # IMEI 추가 후 저장
        # 정합성 대여 단말 리스트에 단말 추가 끝
        else:
            print("이미 등록되어 있는 단말임")

    else:
        print("imei 자리 수 확인 필요")
########################################################################################


########################################################################################
# 단말 반납 처리
def back():
    imei = str(num.get())
    Rimeis = []

    for rx in rentalSheet["F"]:
        Rimeis.append(rx.value) # 대여 리스트 엑셀에서 F열값 리스트 생성

    Floc = (Rimeis.index(str(imei)))+1


    if str(imei) in Rimeis : # IMEI 값이 RIMEIS에 들어있는지 체크
        for x in range(1, rentalSheet.max_column+1):
            # 반납에 해당하는 컬러 값"d9d9d9"
            rentalSheet.cell(row=Floc, column=x).fill = PatternFill(start_color="d9d9d9", end_color="d9d9d9", fill_type="solid")
        rentalSheet.cell(row=Floc, column=12, value="반납")
        rentalSheet.cell(row=Floc, column=13, value="")
        
        rental.save("정합성 단말 대여 리스트_"+ str(today).replace("-","")[2:8] +".xlsx") # IMEI 추가 후 저장
    else:
        print("없는데 뭘 찾는 거야")
########################################################################################


frame2 = Frame(root)
frame2.pack(side="right", fill="both", expand=True)
Button(frame2, text="대여", command=Add).pack(fill="both", expand=True)
Button(frame2, text="반납", command=back).pack(fill="both", expand=True)

root.mainloop()
